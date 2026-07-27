"""Spreadsheet reading for LLM-driven analysis.

``initial_setup._read_xlsx`` reads the *first sheet only*, capped at 100 rows,
which is right for the one-shot setup summary but useless for a model that has
to explore a workbook. This service adds the three operations an agent actually
needs — describe, page through a range, and aggregate — with caps that keep any
single tool result inside the context window.

Aggregation runs in Python on purpose: asking a model to total 5000 rows either
gets the arithmetic wrong or requires shipping the whole sheet into context.
"""

from __future__ import annotations

import csv
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Literal

from negotium.app.initial_setup import _has_sensitive_hint

SPREADSHEET_SUFFIXES = {".xlsx", ".xlsm", ".csv", ".tsv"}
UNSUPPORTED_LEGACY_SUFFIXES = {".xls"}

MAX_SHEETS = 20
MAX_ROWS_PER_READ = 500
MAX_CELLS_PER_READ = 20_000
MAX_CELL_CHARS = 500
MAX_FILE_BYTES = 20 * 1024 * 1024
MAX_GROUPS = 200
_SAMPLE_ROWS = 5

AggregateOp = Literal["sum", "count", "avg", "min", "max"]

# Currency/percent decoration that has to come off before a cell is a number.
_NUMERIC_STRIP_RE = re.compile(r"[,\s₩$€%원건개명]")


@dataclass(frozen=True)
class SheetSummary:
    name: str
    rows: int
    columns: int
    headers: list[str]
    sample: list[dict[str, str]] = field(default_factory=list)
    sensitive_hint: bool = False
    notes: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return {
            "name": self.name,
            "rows": self.rows,
            "columns": self.columns,
            "headers": self.headers,
            "sample": self.sample,
            "sensitive_hint": self.sensitive_hint,
            "notes": self.notes,
        }


@dataclass(frozen=True)
class SheetPage:
    sheet: str
    headers: list[str]
    rows: list[dict[str, str]]
    start_row: int
    end_row: int
    total_rows: int
    truncated: bool = False
    sensitive_hint: bool = False
    notes: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return {
            "sheet": self.sheet,
            "headers": self.headers,
            "rows": self.rows,
            "start_row": self.start_row,
            "end_row": self.end_row,
            "total_rows": self.total_rows,
            "truncated": self.truncated,
            "sensitive_hint": self.sensitive_hint,
            "notes": self.notes,
        }


class UnsupportedSpreadsheetError(ValueError):
    """Raised for formats we deliberately do not support."""


def _check_readable(path: Path) -> None:
    suffix = path.suffix.lower()
    if suffix in UNSUPPORTED_LEGACY_SUFFIXES:
        raise UnsupportedSpreadsheetError(
            "xls(97-2003) 형식은 지원하지 않습니다. xlsx로 다시 저장해 주세요."
        )
    if suffix not in SPREADSHEET_SUFFIXES:
        raise UnsupportedSpreadsheetError(f"표 형식 파일이 아닙니다: {suffix or '(확장자 없음)'}")
    if not path.exists() or not path.is_file():
        raise FileNotFoundError(f"파일을 찾을 수 없습니다: {path.name}")
    if path.stat().st_size > MAX_FILE_BYTES:
        raise UnsupportedSpreadsheetError(
            f"파일이 너무 큽니다({path.stat().st_size // (1024 * 1024)}MB). "
            f"최대 {MAX_FILE_BYTES // (1024 * 1024)}MB까지 읽을 수 있습니다."
        )


def _cell(value: Any) -> str:
    text = "" if value is None else str(value).strip()
    return text[:MAX_CELL_CHARS]


def _headers_from(row: tuple[Any, ...]) -> list[str]:
    headers: list[str] = []
    for index, value in enumerate(row):
        headers.append(_cell(value) or f"col_{index + 1}")
    return headers


def _delimiter_for(path: Path) -> str:
    return "\t" if path.suffix.lower() == ".tsv" else ","


def _read_delimited_rows(path: Path) -> tuple[list[str], list[list[str]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.reader(handle, delimiter=_delimiter_for(path)))
    if not rows:
        return [], []
    headers = _headers_from(tuple(rows[0]))
    return headers, [[_cell(cell) for cell in row] for row in rows[1:]]


def _open_workbook(path: Path) -> Any:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency guard
        raise RuntimeError("XLSX 파싱에는 openpyxl 이 필요합니다.") from exc
    return load_workbook(path, read_only=True, data_only=True)


def _empty_ratio(rows: list[list[str]], columns: int) -> float:
    if not rows or not columns:
        return 0.0
    total = len(rows) * columns
    filled = sum(1 for row in rows for cell in row if cell)
    return 1.0 - (filled / total)


def _formula_note(rows: list[list[str]], columns: int) -> list[str]:
    # With data_only=True openpyxl returns None for formula cells in a workbook
    # Excel has never opened (no cached values). Detecting that precisely is not
    # worth it; a heuristic note is enough for the model to explain itself.
    if _empty_ratio(rows, columns) > 0.3:
        return [
            "빈 셀이 많습니다. 수식 결과가 저장되지 않은 파일일 수 있으니 "
            "Excel에서 한 번 열어 저장한 뒤 다시 업로드해 보세요."
        ]
    return []


def describe_workbook(path: Path) -> list[SheetSummary]:
    """Sheet names, dimensions, headers and a small sample per sheet."""

    _check_readable(path)
    if path.suffix.lower() in {".csv", ".tsv"}:
        headers, rows = _read_delimited_rows(path)
        sample = [_row_to_dict(headers, row) for row in rows[:_SAMPLE_ROWS]]
        return [
            SheetSummary(
                name=path.stem,
                rows=len(rows),
                columns=len(headers),
                headers=headers,
                sample=sample,
                sensitive_hint=_has_sensitive_hint(path.name, " ".join(headers)),
                notes=_formula_note(rows, len(headers)),
            )
        ]

    workbook = _open_workbook(path)
    try:
        summaries: list[SheetSummary] = []
        for name in workbook.sheetnames[:MAX_SHEETS]:
            sheet = workbook[name]
            sheet_headers: list[str] = []
            sample_rows: list[list[str]] = []
            total = 0
            for index, raw in enumerate(sheet.iter_rows(values_only=True)):
                if index == 0:
                    sheet_headers = _headers_from(raw)
                    continue
                total += 1
                if len(sample_rows) < _SAMPLE_ROWS:
                    sample_rows.append([_cell(value) for value in raw])
            summaries.append(
                SheetSummary(
                    name=name,
                    rows=total,
                    columns=len(sheet_headers),
                    headers=sheet_headers,
                    sample=[_row_to_dict(sheet_headers, row) for row in sample_rows],
                    sensitive_hint=_has_sensitive_hint(
                        f"{path.name} {name}", " ".join(sheet_headers)
                    ),
                    notes=_formula_note(sample_rows, len(sheet_headers)),
                )
            )
        if len(workbook.sheetnames) > MAX_SHEETS:
            summaries.append(
                SheetSummary(
                    name="(생략)",
                    rows=0,
                    columns=0,
                    headers=[],
                    notes=[
                        f"시트가 {len(workbook.sheetnames)}개라 앞의 {MAX_SHEETS}개만 표시했습니다."
                    ],
                )
            )
        return summaries
    finally:
        # read_only workbooks hold file handles open; a long-running server
        # leaks descriptors without this.
        workbook.close()


def _row_to_dict(headers: list[str], row: list[str]) -> dict[str, str]:
    return {
        headers[index] if index < len(headers) else f"col_{index + 1}": value
        for index, value in enumerate(row)
    }


def read_sheet(
    path: Path,
    *,
    sheet: str = "",
    start_row: int = 1,
    limit: int = 100,
    columns: list[str] | None = None,
) -> SheetPage:
    """Read a row range. ``start_row`` is 1-based over data rows (header excluded)."""

    _check_readable(path)
    limit = max(1, min(limit, MAX_ROWS_PER_READ))
    start_row = max(1, start_row)
    notes: list[str] = []

    if path.suffix.lower() in {".csv", ".tsv"}:
        headers, all_rows = _read_delimited_rows(path)
        sheet_name = path.stem
        total = len(all_rows)
        window = all_rows[start_row - 1 : start_row - 1 + limit]
    else:
        workbook = _open_workbook(path)
        try:
            sheet_name = sheet or workbook.sheetnames[0]
            if sheet_name not in workbook.sheetnames:
                raise ValueError(
                    f"'{sheet_name}' 시트가 없습니다. 사용 가능: {', '.join(workbook.sheetnames)}"
                )
            worksheet = workbook[sheet_name]
            headers = []
            window = []
            total = 0
            for index, raw in enumerate(worksheet.iter_rows(values_only=True)):
                if index == 0:
                    headers = _headers_from(raw)
                    continue
                total += 1
                if start_row <= total < start_row + limit:
                    window.append([_cell(value) for value in raw])
        finally:
            workbook.close()

    if columns:
        keep = [header for header in headers if header in set(columns)]
        missing = sorted(set(columns) - set(headers))
        if missing:
            notes.append(f"요청한 컬럼을 찾지 못했습니다: {', '.join(missing)}")
        rows = [
            {key: value for key, value in _row_to_dict(headers, row).items() if key in set(keep)}
            for row in window
        ]
        headers = keep or headers
    else:
        rows = [_row_to_dict(headers, row) for row in window]

    truncated = False
    if rows and len(rows) * max(len(headers), 1) > MAX_CELLS_PER_READ:
        allowed = max(1, MAX_CELLS_PER_READ // max(len(headers), 1))
        rows = rows[:allowed]
        truncated = True
        notes.append(f"셀 수 상한({MAX_CELLS_PER_READ})에 맞춰 {allowed}행만 반환했습니다.")
    if start_row - 1 + len(rows) < total:
        truncated = True

    return SheetPage(
        sheet=sheet_name,
        headers=headers,
        rows=rows,
        start_row=start_row,
        end_row=start_row + len(rows) - 1 if rows else start_row - 1,
        total_rows=total,
        truncated=truncated,
        sensitive_hint=_has_sensitive_hint(f"{path.name} {sheet_name}", " ".join(headers)),
        notes=notes,
    )


def to_number(value: str) -> float | None:
    """Parse a spreadsheet cell into a number, tolerating Korean formatting."""

    text = (value or "").strip()
    if not text:
        return None
    negative = text.startswith("(") and text.endswith(")")
    if negative:
        text = text[1:-1]
    text = _NUMERIC_STRIP_RE.sub("", text)
    if not text or text in {"-", "."}:
        return None
    try:
        number = float(text)
    except ValueError:
        return None
    return -number if negative else number


def aggregate_sheet(
    path: Path,
    *,
    sheet: str = "",
    group_by: list[str] | None = None,
    measures: list[str] | None = None,
    agg: AggregateOp = "sum",
) -> dict[str, Any]:
    """Group rows and reduce numeric columns, entirely in Python.

    Reading 5000 rows into the prompt to compute a total is both expensive and
    unreliable; this returns just the grouped result.
    """

    group_by = group_by or []
    measures = measures or []
    notes: list[str] = []

    # Stream the whole sheet in pages so memory stays bounded.
    page = read_sheet(path, sheet=sheet, start_row=1, limit=MAX_ROWS_PER_READ)
    headers = page.headers
    rows: list[dict[str, str]] = list(page.rows)
    cursor = page.end_row + 1
    while cursor <= page.total_rows:
        nxt = read_sheet(path, sheet=sheet, start_row=cursor, limit=MAX_ROWS_PER_READ)
        if not nxt.rows:
            break
        rows.extend(nxt.rows)
        cursor = nxt.end_row + 1

    missing = [column for column in [*group_by, *measures] if column not in headers]
    if missing:
        notes.append(f"존재하지 않는 컬럼을 무시했습니다: {', '.join(missing)}")
        group_by = [column for column in group_by if column in headers]
        measures = [column for column in measures if column in headers]

    buckets: dict[tuple[str, ...], dict[str, list[float]]] = {}
    counts: dict[tuple[str, ...], int] = {}
    for row in rows:
        key = tuple(row.get(column, "") for column in group_by)
        counts[key] = counts.get(key, 0) + 1
        bucket = buckets.setdefault(key, {measure: [] for measure in measures})
        for measure in measures:
            number = to_number(row.get(measure, ""))
            if number is not None:
                bucket[measure].append(number)

    groups: list[dict[str, Any]] = []
    for key, bucket in list(buckets.items())[:MAX_GROUPS]:
        entry: dict[str, Any] = {column: key[index] for index, column in enumerate(group_by)}
        entry["count"] = counts[key]
        for measure, values in bucket.items():
            entry[measure] = _reduce(values, agg)
        groups.append(entry)
    if len(buckets) > MAX_GROUPS:
        notes.append(f"그룹이 {len(buckets)}개라 상위 {MAX_GROUPS}개만 반환했습니다.")

    return {
        "sheet": page.sheet,
        "agg": agg,
        "group_by": group_by,
        "measures": measures,
        "total_rows": page.total_rows,
        "groups": groups,
        "sensitive_hint": page.sensitive_hint,
        "notes": notes,
    }


def _reduce(values: list[float], agg: AggregateOp) -> float | int | None:
    if agg == "count":
        return len(values)
    if not values:
        return None
    if agg == "sum":
        return round(sum(values), 6)
    if agg == "avg":
        return round(sum(values) / len(values), 6)
    if agg == "min":
        return min(values)
    return max(values)
