"""Multi-sheet spreadsheet reading, caps, and Korean-formatted aggregation."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from negotium.app.services.spreadsheet_service import (
    MAX_ROWS_PER_READ,
    UnsupportedSpreadsheetError,
    aggregate_sheet,
    describe_workbook,
    read_sheet,
    to_number,
)


def _workbook(path: Path) -> Path:
    book = Workbook()
    sales = book.active
    sales.title = "매출"
    sales.append(["부서", "담당자", "금액", "건수"])
    sales.append(["영업1팀", "김민수", "1,200,000", "3"])
    sales.append(["영업1팀", "이서연", "₩800,000", "2"])
    sales.append(["영업2팀", "박지훈", "2,000,000", "5"])
    sales.append(["영업2팀", "최유진", "(500,000)", "1"])

    staff = book.create_sheet("인원")
    staff.append(["이름", "직급"])
    staff.append(["김민수", "과장"])

    book.save(path)
    return path


def test_describe_reports_every_sheet(tmp_path: Path) -> None:
    """The old parser only ever saw ``workbook.active``."""

    path = _workbook(tmp_path / "sales.xlsx")
    sheets = {sheet.name: sheet for sheet in describe_workbook(path)}

    assert set(sheets) == {"매출", "인원"}
    assert sheets["매출"].rows == 4
    assert sheets["매출"].headers == ["부서", "담당자", "금액", "건수"]
    assert sheets["매출"].sample[0]["담당자"] == "김민수"
    assert sheets["인원"].rows == 1


def test_read_sheet_pages_through_a_named_sheet(tmp_path: Path) -> None:
    path = _workbook(tmp_path / "sales.xlsx")

    page = read_sheet(path, sheet="매출", start_row=2, limit=2)

    assert page.sheet == "매출"
    assert page.total_rows == 4
    assert page.start_row == 2
    assert page.end_row == 3
    assert [row["담당자"] for row in page.rows] == ["이서연", "박지훈"]
    assert page.truncated is True  # rows remain after this window


def test_read_sheet_column_projection_and_missing_column_note(tmp_path: Path) -> None:
    path = _workbook(tmp_path / "sales.xlsx")

    page = read_sheet(path, sheet="매출", columns=["담당자", "없는컬럼"])

    assert set(page.rows[0]) == {"담당자"}
    assert any("없는컬럼" in note for note in page.notes)


def test_unknown_sheet_name_lists_the_options(tmp_path: Path) -> None:
    path = _workbook(tmp_path / "sales.xlsx")

    with pytest.raises(ValueError) as excinfo:
        read_sheet(path, sheet="없는시트")

    assert "매출" in str(excinfo.value)


def test_row_limit_is_capped(tmp_path: Path) -> None:
    book = Workbook()
    sheet = book.active
    sheet.append(["n"])
    for index in range(MAX_ROWS_PER_READ + 50):
        sheet.append([index])
    path = tmp_path / "big.xlsx"
    book.save(path)

    page = read_sheet(path, limit=10_000)

    assert len(page.rows) == MAX_ROWS_PER_READ
    assert page.truncated is True


@pytest.mark.parametrize(
    ("raw", "expected"),
    [
        ("1,200,000", 1_200_000.0),
        ("₩800,000", 800_000.0),
        ("12%", 12.0),
        ("3건", 3.0),
        ("(500,000)", -500_000.0),
        ("", None),
        ("해당없음", None),
    ],
)
def test_korean_number_formatting_is_parsed(raw: str, expected: float | None) -> None:
    assert to_number(raw) == expected


def test_aggregate_groups_and_sums_in_python(tmp_path: Path) -> None:
    """Totals are computed in code, not by asking the model to add up rows."""

    path = _workbook(tmp_path / "sales.xlsx")

    result = aggregate_sheet(
        path, sheet="매출", group_by=["부서"], measures=["금액", "건수"], agg="sum"
    )

    groups = {group["부서"]: group for group in result["groups"]}
    assert groups["영업1팀"]["금액"] == 2_000_000
    assert groups["영업1팀"]["건수"] == 5
    assert groups["영업1팀"]["count"] == 2
    # Parenthesised values are negative in accounting exports.
    assert groups["영업2팀"]["금액"] == 1_500_000


def test_aggregate_avg_and_missing_column_note(tmp_path: Path) -> None:
    path = _workbook(tmp_path / "sales.xlsx")

    result = aggregate_sheet(
        path, sheet="매출", group_by=["부서"], measures=["금액", "없음"], agg="avg"
    )

    assert any("없음" in note for note in result["notes"])
    groups = {group["부서"]: group for group in result["groups"]}
    assert groups["영업1팀"]["금액"] == 1_000_000


def test_csv_is_read_through_the_same_api(tmp_path: Path) -> None:
    path = tmp_path / "list.csv"
    path.write_text("이름,부서\n김민수,영업1팀\n이서연,영업2팀\n", encoding="utf-8")

    sheets = describe_workbook(path)
    page = read_sheet(path)

    assert sheets[0].rows == 2
    assert [row["이름"] for row in page.rows] == ["김민수", "이서연"]


def test_legacy_xls_is_refused_with_guidance(tmp_path: Path) -> None:
    path = tmp_path / "old.xls"
    path.write_bytes(b"\xd0\xcf\x11\xe0")

    with pytest.raises(UnsupportedSpreadsheetError) as excinfo:
        describe_workbook(path)

    assert "xlsx" in str(excinfo.value)


def test_sensitive_headers_are_flagged(tmp_path: Path) -> None:
    book = Workbook()
    sheet = book.active
    sheet.title = "급여대장"
    sheet.append(["이름", "급여", "주민등록번호"])
    sheet.append(["김민수", "5000000", "900101-1234567"])
    path = tmp_path / "payroll.xlsx"
    book.save(path)

    summary = describe_workbook(path)[0]
    page = read_sheet(path)

    assert summary.sensitive_hint is True
    assert page.sensitive_hint is True
