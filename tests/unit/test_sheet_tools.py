"""MCP sheet/org tools: path safety, sensitive-data gating, org reads."""

from __future__ import annotations

import io
import json
from pathlib import Path

import pytest
from openpyxl import Workbook

from negotium.app.container import Container
from negotium.app.services.mcp_hub_service import call_tool, is_read_tool, list_tool_descriptors
from negotium.app.settings import Settings
from negotium.archive.access_control import DepartmentRecord, PositionRecord, UserRecord
from negotium.archive.llm_runtime import LlmRuntimeConfig


def _container(tmp_path: Path, *, local_enabled: bool = False) -> Container:
    container = Container.build(
        Settings(
            env="test",
            archive_dir=tmp_path / "archive",
            workspace_dir=tmp_path / "workspaces",
        )
    )
    container.llm_runtime.write(
        LlmRuntimeConfig(
            local_enabled=local_enabled,
            api_enabled=True,
            default_route="api",
            default_provider="solar",
        )
    )
    return container


def _upload_workbook(container: Container, name: str, build) -> str:
    book = Workbook()
    build(book)
    buffer = io.BytesIO()
    book.save(buffer)
    buffer.seek(0)
    return container.uploads.save(filename=name, source=buffer).id


def _sales(book: Workbook) -> None:
    sheet = book.active
    sheet.title = "매출"
    sheet.append(["부서", "금액"])
    sheet.append(["영업1팀", "1,000"])
    sheet.append(["영업2팀", "2,000"])


def _payroll(book: Workbook) -> None:
    sheet = book.active
    sheet.title = "급여"
    sheet.append(["이름", "급여", "주민등록번호"])
    sheet.append(["김민수", "5000000", "900101-1234567"])


def test_sheet_and_org_tools_are_registered_as_read_tools() -> None:
    names = {descriptor["name"] for descriptor in list_tool_descriptors()}
    expected = {
        "sheets.list_files",
        "sheets.describe",
        "sheets.read_range",
        "sheets.aggregate",
        "org.roster",
        "org.find_people",
    }
    assert expected <= names
    # Read tools auto-execute in the agent loop; a misclassification here would
    # either block them behind approval or let a writer run unapproved.
    for name in expected:
        assert is_read_tool(name) is True


def test_describe_and_read_range_round_trip(tmp_path: Path) -> None:
    container = _container(tmp_path)
    upload_id = _upload_workbook(container, "sales.xlsx", _sales)

    described = call_tool(container, "sheets.describe", {"upload_id": upload_id})
    assert described.result["sheets"][0]["name"] == "매출"

    page = call_tool(container, "sheets.read_range", {"upload_id": upload_id, "sheet": "매출"})
    assert [row["부서"] for row in page.result["rows"]] == ["영업1팀", "영업2팀"]


def test_aggregate_returns_computed_totals(tmp_path: Path) -> None:
    container = _container(tmp_path)
    upload_id = _upload_workbook(container, "sales.xlsx", _sales)

    result = call_tool(
        container,
        "sheets.aggregate",
        {"upload_id": upload_id, "sheet": "매출", "group_by": ["부서"], "measures": ["금액"]},
    )

    totals = {group["부서"]: group["금액"] for group in result.result["groups"]}
    assert totals == {"영업1팀": 1000, "영업2팀": 2000}


def test_list_files_only_returns_spreadsheets(tmp_path: Path) -> None:
    container = _container(tmp_path)
    _upload_workbook(container, "sales.xlsx", _sales)
    container.uploads.save(filename="notes.txt", source=io.BytesIO(b"hello"))

    result = call_tool(container, "sheets.list_files", {})

    assert [item["filename"] for item in result.result["files"]] == ["sales.xlsx"]


def test_sensitive_sheet_is_redacted_on_a_cloud_route(tmp_path: Path) -> None:
    """Raw 급여/주민등록 cells must not reach an external model.

    Tool results are fed straight back into the next completion, so on a cloud
    route the values would leave the building.
    """

    container = _container(tmp_path, local_enabled=True)
    upload_id = _upload_workbook(container, "payroll.xlsx", _payroll)

    result = call_tool(container, "sheets.read_range", {"upload_id": upload_id}, route="cloud")

    assert result.result.get("redacted") is True
    assert "rows" not in result.result
    assert result.result["headers"] == ["이름", "급여", "주민등록번호"]
    assert result.result["redacted_row_count"] == 1
    assert any("민감정보" in note for note in result.result["notes"])
    # The actual resident-registration number must appear nowhere in the payload.
    assert "900101-1234567" not in json.dumps(result.result, ensure_ascii=False)


def test_sensitive_sheet_is_returned_on_the_local_route(tmp_path: Path) -> None:
    """Nothing leaves the building on a local route, so no redaction is needed."""

    container = _container(tmp_path, local_enabled=True)
    upload_id = _upload_workbook(container, "payroll.xlsx", _payroll)

    result = call_tool(container, "sheets.read_range", {"upload_id": upload_id}, route="local")

    assert result.result.get("redacted") is not True
    assert result.result["rows"]


def test_gate_defaults_to_redacting_when_the_route_is_unknown(tmp_path: Path) -> None:
    """local_enabled defaults to True, so keying on it disabled the gate entirely.

    The policy has to follow the route of the actual call, and an unknown route
    must fail safe rather than fail open.
    """

    container = _container(tmp_path, local_enabled=True)
    upload_id = _upload_workbook(container, "payroll.xlsx", _payroll)

    result = call_tool(container, "sheets.read_range", {"upload_id": upload_id})

    assert result.result.get("redacted") is True
    assert "900101-1234567" not in json.dumps(result.result, ensure_ascii=False)


def test_describe_does_not_crash_on_a_sensitive_sheet(tmp_path: Path) -> None:
    """SheetSummary.rows is an int row-count, not a list — len() blew up."""

    container = _container(tmp_path, local_enabled=True)
    upload_id = _upload_workbook(container, "payroll.xlsx", _payroll)

    result = call_tool(container, "sheets.describe", {"upload_id": upload_id}, route="cloud")

    sheet = result.result["sheets"][0]
    assert sheet["redacted"] is True
    assert sheet["redacted_row_count"] == 1
    assert "sample" not in sheet
    assert "900101-1234567" not in json.dumps(result.result, ensure_ascii=False)


def test_aggregate_masks_sensitive_group_keys(tmp_path: Path) -> None:
    """group_by keys are verbatim cell values, so they need masking too.

    Without this, `group_by=["주민등록번호"]` returned raw PII straight past the
    read_range gate.
    """

    container = _container(tmp_path, local_enabled=True)
    upload_id = _upload_workbook(container, "payroll.xlsx", _payroll)

    result = call_tool(
        container,
        "sheets.aggregate",
        {"upload_id": upload_id, "group_by": ["주민등록번호"], "measures": ["급여"]},
        route="cloud",
    )

    assert "900101-1234567" not in json.dumps(result.result, ensure_ascii=False)
    assert result.result["redacted_columns"] == ["주민등록번호"]
    # The derived number is still usable.
    assert result.result["groups"][0]["급여"] == 5_000_000


def test_aggregate_keeps_non_sensitive_group_keys(tmp_path: Path) -> None:
    container = _container(tmp_path, local_enabled=True)
    upload_id = _upload_workbook(container, "sales.xlsx", _sales)

    result = call_tool(
        container,
        "sheets.aggregate",
        {"upload_id": upload_id, "sheet": "매출", "group_by": ["부서"], "measures": ["금액"]},
        route="cloud",
    )

    totals = {group["부서"]: group["금액"] for group in result.result["groups"]}
    assert totals == {"영업1팀": 1000, "영업2팀": 2000}


def test_aggregates_survive_the_sensitive_gate(tmp_path: Path) -> None:
    """Derived numbers are safe even when raw cells are withheld."""

    container = _container(tmp_path, local_enabled=False)
    upload_id = _upload_workbook(container, "payroll.xlsx", _payroll)

    result = call_tool(
        container,
        "sheets.aggregate",
        {"upload_id": upload_id, "group_by": [], "measures": ["급여"]},
        route="cloud",
    )

    assert result.result["groups"][0]["급여"] == 5_000_000


def test_unknown_upload_id_raises_a_clean_error(tmp_path: Path) -> None:
    container = _container(tmp_path)

    with pytest.raises(FileNotFoundError):
        call_tool(container, "sheets.describe", {"upload_id": "does-not-exist"})


def test_missing_upload_id_is_reported(tmp_path: Path) -> None:
    container = _container(tmp_path)

    with pytest.raises(ValueError, match="upload_id"):
        call_tool(container, "sheets.describe", {})


def test_index_path_traversal_is_refused(tmp_path: Path) -> None:
    """The uploads index is a plain JSON file and is treated as untrusted."""

    container = _container(tmp_path)
    record_id = _upload_workbook(container, "sales.xlsx", _sales)
    index = tmp_path / "archive" / "uploads" / "index.json"
    payload = json.loads(index.read_text(encoding="utf-8"))
    payload[0]["path"] = "../../../../etc/passwd"
    index.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")

    with pytest.raises(ValueError, match="archive 외부"):
        container.uploads.resolve_path(record_id)


def test_org_tools_read_the_roster(tmp_path: Path) -> None:
    container = _container(tmp_path)
    container.access_control.upsert_department(
        DepartmentRecord(id="sales", name="영업팀", description="", lead_user_id="", parent_id="")
    )
    container.access_control.upsert_position(
        PositionRecord(id="mgr", name="과장", permissions=[], display_order=1, level=3)
    )
    container.access_control.upsert_user(
        UserRecord(
            id="u1",
            display_name="김민수",
            title="영업 담당",
            role_id="staff",
            department="sales",
            position_id="mgr",
        )
    )

    roster = call_tool(container, "org.roster", {})
    assert roster.result["total_users"] == 1
    assert roster.result["users"][0]["department"] == "영업팀"
    assert roster.result["users"][0]["position"] == "과장"

    markdown = call_tool(container, "org.roster", {"format": "markdown"})
    assert "영업팀" in markdown.result["markdown"]

    found = call_tool(container, "org.find_people", {"query": "김민수"})
    assert [user["display_name"] for user in found.result["matches"]] == ["김민수"]

    missing = call_tool(container, "org.find_people", {"query": "없는사람"})
    assert missing.result["matches"] == []
